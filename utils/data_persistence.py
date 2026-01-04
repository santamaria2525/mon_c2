"""
Data persistence utilities for Monster Strike Bot

This module handles data storage operations including:
- CSV file operations (read/write)
- Excel file operations with image embedding
- Data validation and error handling
"""

import os
import csv
import time
import threading
from datetime import datetime
from typing import List, Tuple, Optional
from logging_util import logger
from openpyxl import Workbook, load_workbook
from openpyxl.drawing import image
from openpyxl.styles import Alignment

# Windows環境ではthreadingによる排他制御のみ使用

def update_csv_data(filename: str, folder: str, orbs: int, found_character: bool, 
                   account_name: str = None, account_image: str = None, 
                   orb_image: str = None) -> bool:
    """
    CSVファイルにデータを追加する
    
    Args:
        filename: CSVファイル名
        folder: フォルダ名
        orbs: オーブ数
        found_character: キャラクターを見つけたかどうか
        account_name: アカウント名（オプション）
        account_image: アカウント名画像パス（オプション）
        orb_image: オーブ数画像パス（オプション）
        
    Returns:
        bool: 更新成功かどうか
    """
    try:
        # ディレクトリが存在しない場合は作成
        os.makedirs(os.path.dirname(os.path.abspath(filename)) if os.path.dirname(filename) else '.', exist_ok=True)
        
        file_exists = os.path.exists(filename)
        mode = 'a' if file_exists else 'w'

        with open(filename, mode, newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)

            # 新規ファイルの場合はヘッダーを書き込む
            if not file_exists:
                csv_writer.writerow(["Account Name", "Account Image", "Folder", "Orb Count", "Orb Image", "Found Character", "Timestamp"])

            # データと現在時刻を書き込む
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            csv_writer.writerow([account_name or "Unknown", account_image or "", folder, orbs, orb_image or "", found_character, timestamp])

        return True
        
    except Exception as e:
        logger.error(f"CSVファイルへの記録中にエラーが発生しました: {str(e)}")
        return False

def read_csv_data(filename: str) -> List[Tuple[str, int]]:
    """
    CSVファイルからデータを読み込む
    
    Args:
        filename: CSVファイル名
        
    Returns:
        List[Tuple[str, int]]: [フォルダ名, オーブ数] のリスト
    """
    data = []
    
    # ファイルが存在しない場合は空リストを返す
    if not os.path.exists(filename):
        logger.warning(f"CSVファイルが存在しません: {filename}")
        return data
        
    try:
        with open(filename, 'r', newline='', encoding='utf-8') as csvfile:
            csv_reader = csv.reader(csvfile)
            
            # ヘッダー行をスキップ
            next(csv_reader, None)
            
            # 各行を処理
            for row in csv_reader:
                if len(row) >= 2:  # 少なくともフォルダ名とオーブ数が必要
                    folder_name, orbs = row[0], row[1]
                    try:
                        # オーブ数を整数に変換
                        data.append((folder_name, int(orbs)))
                    except ValueError:
                        logger.warning(f"無効なオーブ数: {orbs} (フォルダ: {folder_name})")
    except Exception as e:
        logger.error(f"CSVファイルの読み込み中にエラーが発生しました: {str(e)}")
    
    return data

# グローバルロックオブジェクト（エクセルファイルごと）
_excel_locks = {}
_lock_creation_lock = threading.Lock()

def _get_excel_lock(filepath: str) -> threading.RLock:
    """エクセルファイルごとの排他制御ロックを取得"""
    with _lock_creation_lock:
        if filepath not in _excel_locks:
            _excel_locks[filepath] = threading.RLock()
        return _excel_locks[filepath]

def update_excel_data(filename: str, folder: str, orbs: int, found_character: bool, 
                     account_name: str = None, account_image: str = None, 
                     orb_image: str = None, character_ownership_image: str = None,
                     player_id: str = None) -> bool:
    """
    Excelファイルにデータを追加し、画像をセルに貼り付ける（完全排他制御版）
    
    Args:
        filename: Excelファイル名
        folder: フォルダ名
        orbs: オーブ数
        found_character: キャラクターを見つけたかどうか
        account_name: アカウント名（オプション）
        account_image: アカウント名画像パス（オプション）
        orb_image: オーブ数画像パス（オプション）
        character_ownership_image: キャラ所持確認画像パス（オプション）
        player_id: プレイヤーID（オプション）
        
    Returns:
        bool: 更新成功かどうか
    """
    import tempfile
    import shutil
    import msvcrt
    
    # エクセルファイルパスを確定
    excel_filename = filename if filename.endswith('.xlsx') else filename.replace('.csv', '.xlsx')
    
    # exe階層にファイルを保存（大改修前の動作を再現）
    from utils.path_manager import get_base_path
    base_path = get_base_path()
    excel_filepath = os.path.join(base_path, excel_filename)
    
    # ファイルごとの排他制御ロック取得
    file_lock = _get_excel_lock(excel_filepath)
    
    with file_lock:  # 完全排他制御開始
        try:
            logger.info(f"エクセルファイル保存開始: {excel_filepath} (フォルダ: {folder})")
            
            # ディレクトリ作成（大改修前の方法）
            os.makedirs(os.path.dirname(os.path.abspath(excel_filepath)) if os.path.dirname(excel_filepath) else '.', exist_ok=True)
        
            # 一時ファイルを使用した安全な書き込み
            temp_dir = os.path.dirname(excel_filepath) if os.path.dirname(excel_filepath) else '.'
            temp_file = None
            
            # 既存ファイルを開くか新規作成（排他制御下で安全に実行）
            if os.path.exists(excel_filepath):
                wb = load_workbook(excel_filepath)
                ws = wb.active
                next_row = ws.max_row + 1
                logger.debug(f"既存エクセル読み込み完了。次の行: {next_row}")
            else:
                wb = Workbook()
                ws = wb.active
                # ヘッダーを作成
                headers = ["Account Name", "Account Image", "Folder", "Orb Count", "Orb Image", "Found Character", "Character Ownership Image", "Player ID", "Timestamp"]
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)
                next_row = 2
                logger.debug("新規エクセル作成。次の行: 2")
        
            # データを追加
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row_data = [account_name or "Unknown", "", folder, orbs, "", found_character, "", player_id or "", timestamp]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col_num, value=value)
            
            logger.debug(f"行データ追加完了: 行{next_row}, フォルダ{folder}")
            
            # 行の高さを設定（画像表示用）
            ws.row_dimensions[next_row].height = 60
        
            # 画像を貼り付け
            if account_image:
                account_image_path = os.path.abspath(account_image) if not os.path.isabs(account_image) else account_image
                if os.path.exists(account_image_path):
                    try:
                        img = image.Image(account_image_path)
                        # 画像サイズを調整（より大きく）
                        img.width = 120
                        img.height = 50
                        # B列（Account Image）に画像を配置
                        cell_ref = f'B{next_row}'
                        ws.add_image(img, cell_ref)
                        # Account Imageセルのテキストを削除
                        ws[cell_ref].value = None
                    except Exception as e:
                        logger.error(f"アカウント画像の貼り付けに失敗: {e}")
                        # 失敗時はファイルパスを記録
                        ws.cell(row=next_row, column=2, value=f"ERROR: {account_image}")
                else:
                    logger.warning(f"アカウント画像ファイルが存在しません: {account_image_path}")
                    ws.cell(row=next_row, column=2, value=f"NOT_FOUND: {account_image}")
            
            if orb_image:
                orb_image_path = os.path.abspath(orb_image) if not os.path.isabs(orb_image) else orb_image
                if os.path.exists(orb_image_path):
                    try:
                        img = image.Image(orb_image_path)
                        # 画像サイズを調整（より大きく）
                        img.width = 120
                        img.height = 50
                        # E列（Orb Image）に画像を配置
                        cell_ref = f'E{next_row}'
                        ws.add_image(img, cell_ref)
                        # Orb Imageセルのテキストを削除
                        ws[cell_ref].value = None
                    except Exception as e:
                        logger.error(f"オーブ画像の貼り付けに失敗: {e}")
                        # 失敗時はファイルパスを記録
                        ws.cell(row=next_row, column=5, value=f"ERROR: {orb_image}")
                else:
                    logger.warning(f"オーブ画像ファイルが存在しません: {orb_image_path}")
                    ws.cell(row=next_row, column=5, value=f"NOT_FOUND: {orb_image}")
            
            # キャラ所持確認画像を貼り付け
            if character_ownership_image:
                char_image_path = os.path.abspath(character_ownership_image) if not os.path.isabs(character_ownership_image) else character_ownership_image
                
                if os.path.exists(char_image_path):
                    try:
                        img = image.Image(char_image_path)
                        # 画像サイズを調整
                        img.width = 180
                        img.height = 60
                        # G列（Character Ownership Image）に画像を配置
                        cell_ref = f'G{next_row}'
                        ws.add_image(img, cell_ref)
                        # Character Ownership Imageセルのテキストを削除
                        ws[cell_ref].value = None
                    except Exception as e:
                        logger.error(f"キャラ所持確認画像の貼り付けに失敗: {e}")
                        logger.error(f"画像パス: {char_image_path}")
                        logger.error(f"エラー詳細: {type(e).__name__}: {str(e)}")
                        # 失敗時はファイルパスを記録
                        ws.cell(row=next_row, column=7, value=f"ERROR: {str(e)}")
                else:
                    logger.warning(f"キャラ所持確認画像ファイルが存在しません: {char_image_path}")
                    ws.cell(row=next_row, column=7, value=f"NOT_FOUND: {char_image_path}")
            
            # 列幅を調整（画像表示に適したサイズ）
            ws.column_dimensions['A'].width = 15  # Account Name
            ws.column_dimensions['B'].width = 20  # Account Image（画像用に拡張）
            ws.column_dimensions['C'].width = 10  # Folder
            ws.column_dimensions['D'].width = 12  # Orb Count
            ws.column_dimensions['E'].width = 20  # Orb Image（画像用に拡張）
            ws.column_dimensions['F'].width = 15  # Found Character
            ws.column_dimensions['G'].width = 25  # Character Ownership Image（画像用に拡張）
            ws.column_dimensions['H'].width = 20  # Timestamp
            
            # セルの配置を中央揃えに
            for col in range(1, 9):
                ws.cell(row=next_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            # ファイルを保存（一時ファイルを使用した安全な書き込み）
            save_attempts = 5
            save_wait_time = 0.3
            
            for save_attempt in range(save_attempts):
                try:
                    # 一時ファイルに保存
                    temp_file = tempfile.NamedTemporaryFile(mode='wb', delete=False, dir=temp_dir, suffix='.xlsx')
                    wb.save(temp_file.name)
                    temp_file.close()
                    
                    # 一時ファイルを最終ファイルに移動（アトミックな操作）
                    backup_file = None
                    if os.path.exists(excel_filepath):
                        backup_file = excel_filepath + f'.backup_{int(time.time())}'
                        shutil.move(excel_filepath, backup_file)
                    
                    shutil.move(temp_file.name, excel_filepath)
                    
                    # バックアップファイルを削除
                    if backup_file and os.path.exists(backup_file):
                        os.remove(backup_file)
                    
                    # 保存確認: ファイルが正常に保存されたかチェック
                    if not _verify_excel_save(excel_filepath, next_row, folder):
                        raise Exception("保存後の検証に失敗")
                        
                    logger.info(f"エクセルファイル保存完了: {excel_filepath} (行{next_row}, フォルダ{folder})")
                    return True
                    
                except Exception as save_error:
                    logger.debug(f"Excelファイル保存試行{save_attempt + 1}/{save_attempts}: {save_error}")
                    
                    # 一時ファイルのクリーンアップ
                    if temp_file and os.path.exists(temp_file.name):
                        try:
                            os.remove(temp_file.name)
                        except:
                            pass
                    
                    # バックアップファイルの復元
                    if 'backup_file' in locals() and backup_file and os.path.exists(backup_file):
                        try:
                            shutil.move(backup_file, excel_filepath)
                        except:
                            pass
                    
                    if save_attempt < save_attempts - 1:
                        time.sleep(save_wait_time)
                        save_wait_time *= 1.5
                    else:
                        raise save_error
            
        except Exception as e:
            logger.error(f"Excelファイルへの記録中にエラーが発生しました: {str(e)}")
            logger.error(f"エラー詳細: {type(e).__name__}: {str(e)}")
            logger.error(f"保存先パス: {excel_filepath if 'excel_filepath' in locals() else 'Unknown'}")
            return False

def _verify_excel_save(filepath: str, expected_row: int, expected_folder: str) -> bool:
    """
    エクセルファイルが正常に保存されたかを検証
    
    Args:
        filepath: エクセルファイルパス
        expected_row: 期待される行番号
        expected_folder: 期待されるフォルダ名
    
    Returns:
        bool: 検証成功かどうか
    """
    try:
        if not os.path.exists(filepath):
            logger.error(f"保存確認: ファイルが存在しません: {filepath}")
            return False
        
        # ファイルを再読み込みして確認
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 行数チェック
        if ws.max_row < expected_row:
            logger.error(f"保存確認: 行数不足 期待{expected_row} 実際{ws.max_row}")
            return False
        
        # 該当行のフォルダ名チェック
        saved_folder = ws.cell(row=expected_row, column=3).value
        if str(saved_folder) != str(expected_folder):
            logger.error(f"保存確認: フォルダ名不一致 期待{expected_folder} 実際{saved_folder}")
            return False
        
        logger.debug(f"保存確認成功: 行{expected_row}, フォルダ{expected_folder}")
        return True
        
    except Exception as e:
        logger.error(f"保存確認中にエラー: {e}")
        return False


def update_orb_player_id(filename: str, folder: str, player_id: str) -> bool:
    """
    既存のオーブデータExcelにプレイヤーIDを追記する。

    Args:
        filename: Excelファイル名（例: orb_data.xlsx）
        folder: フォルダ番号
        player_id: 取得したプレイヤーID

    Returns:
        bool: 更新成功かどうか
    """
    if not player_id:
        logger.warning("Player IDが空のためExcel更新をスキップします")
        return False

    excel_filename = filename if filename.endswith(".xlsx") else filename.replace(".csv", ".xlsx")
    from utils.path_manager import get_base_path
    base_path = get_base_path()
    excel_filepath = os.path.join(base_path, excel_filename)

    if not os.path.exists(excel_filepath):
        logger.warning("orbデータExcelが存在しません: %s", excel_filepath)
        return False

    file_lock = _get_excel_lock(excel_filepath)
    with file_lock:
        try:
            wb = load_workbook(excel_filepath)
            ws = wb.active
            target_row = None
            normalized_target = str(folder).zfill(3)

            for row in range(ws.max_row, 1, -1):
                cell_value = ws.cell(row=row, column=3).value
                if cell_value is None:
                    continue
                normalized_cell = str(cell_value).zfill(3)
                if normalized_cell == normalized_target:
                    target_row = row
                    break

            if target_row is None:
                logger.warning("フォルダ%sに対応する行が見つからないためID更新をスキップします", folder)
                return False

            ws.cell(row=target_row, column=8, value=player_id)

            import tempfile
            import shutil

            temp_dir = os.path.dirname(excel_filepath) if os.path.dirname(excel_filepath) else "."
            with tempfile.NamedTemporaryFile(mode="wb", delete=False, dir=temp_dir, suffix=".xlsx") as temp_file:
                wb.save(temp_file.name)
                temp_filepath = temp_file.name

            shutil.move(temp_filepath, excel_filepath)
            logger.info("orbデータExcelにPlayer IDを追記しました (フォルダ=%s, ID=%s)", folder, player_id)
            return True
        except Exception as exc:
            logger.error("orbデータExcelのID更新でエラー: %s", exc, exc_info=True)
            return False
