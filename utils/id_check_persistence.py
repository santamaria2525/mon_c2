"""
utils.id_check_persistence - ID確認専用のエクセル保存機能

IDチェック処理で取得した画像をid_check.xlsxに保存する機能を提供します。
"""

from __future__ import annotations

import os
import threading
import time
from typing import Dict, Optional
from pathlib import Path

from logging_util import logger

# ファイル排他制御用のロック辞書
_excel_locks: Dict[str, threading.RLock] = {}

def _get_excel_lock(filepath: str) -> threading.RLock:
    """ファイル毎の排他制御ロックを取得"""
    if filepath not in _excel_locks:
        _excel_locks[filepath] = threading.RLock()
    return _excel_locks[filepath]

def save_id_check_data(folder: str, id_image_path: str) -> bool:
    """
    ID確認データをid_check.xlsxに保存します。
    
    Args:
        folder: フォルダ名
        id_image_path: ID部分の画像パス
        
    Returns:
        bool: 保存成功かどうか
    """
    import openpyxl
    try:
        from openpyxl.drawing.image import Image as ExcelImage
    except ImportError:
        try:
            from openpyxl.drawing import Image as ExcelImage
        except ImportError:
            logger.error("openpyxlの画像機能が利用できません。画像埋め込みはスキップされます。")
            ExcelImage = None
    import tempfile
    import shutil
    
    # エクセルファイルパスを確定
    from utils.path_manager import get_base_path
    base_path = get_base_path()
    excel_filepath = os.path.join(base_path, "id_check.xlsx")
    
    # ファイルごとの排他制御ロック取得
    file_lock = _get_excel_lock(excel_filepath)
    
    with file_lock:  # 完全排他制御開始
        try:
            logger.info(f"ID確認エクセルファイル保存開始: {excel_filepath} (フォルダ: {folder})")
            
            # ディレクトリ作成
            os.makedirs(os.path.dirname(os.path.abspath(excel_filepath)) if os.path.dirname(excel_filepath) else '.', exist_ok=True)
        
            # 一時ファイルを使用した安全な書き込み
            temp_dir = os.path.dirname(excel_filepath) if os.path.dirname(excel_filepath) else '.'
            temp_file = None

            try:
                # エクセルファイルの読み込みまたは新規作成
                if os.path.exists(excel_filepath):
                    workbook = openpyxl.load_workbook(excel_filepath)
                    worksheet = workbook.active
                    logger.info("既存のid_check.xlsxファイルを読み込みました")
                else:
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    # ヘッダー作成
                    worksheet['A1'] = 'フォルダ名'
                    worksheet['B1'] = 'ID画像'
                    worksheet['C1'] = '処理日時'
                    # 列幅設定
                    worksheet.column_dimensions['A'].width = 15
                    worksheet.column_dimensions['B'].width = 30
                    worksheet.column_dimensions['C'].width = 20
                    logger.info("新しいid_check.xlsxファイルを作成しました")

                # 次の行番号を取得
                next_row = worksheet.max_row + 1
                
                # データを書き込み
                worksheet[f'A{next_row}'] = folder
                worksheet[f'C{next_row}'] = time.strftime("%Y-%m-%d %H:%M:%S")
                
                # ID画像を埋め込み
                if id_image_path and os.path.exists(id_image_path) and ExcelImage is not None:
                    try:
                        # 画像をエクセルに追加
                        img = ExcelImage(id_image_path)
                        # 画像サイズを調整（幅150px, 高さ30px程度）
                        img.width = 150
                        img.height = 30
                        # セルB列に画像を配置
                        img.anchor = f'B{next_row}'
                        worksheet.add_image(img)
                        
                        # 行の高さを画像に合わせて調整
                        worksheet.row_dimensions[next_row].height = 25
                        
                        logger.info(f"ID画像をエクセルに埋め込み: {id_image_path}")
                    except Exception as e:
                        logger.error(f"ID画像の埋め込みに失敗: {e}")
                        worksheet[f'B{next_row}'] = f"画像読み込み失敗: {os.path.basename(id_image_path)}"
                else:
                    # 画像埋め込みが利用できない場合、ファイルパスを記録
                    if id_image_path:
                        worksheet[f'B{next_row}'] = os.path.basename(id_image_path)
                        logger.info(f"ID画像パスを記録: {id_image_path}")
                    else:
                        worksheet[f'B{next_row}'] = "画像なし"
                        logger.warning("ID画像パスが無効です")
                
                # 一時ファイルに保存
                with tempfile.NamedTemporaryFile(mode='wb', delete=False, dir=temp_dir, suffix='.xlsx') as temp_file:
                    workbook.save(temp_file.name)
                    temp_filepath = temp_file.name
                
                # 一時ファイルを本来のファイルに移動（原子的操作）
                shutil.move(temp_filepath, excel_filepath)
                
                logger.info(f"✅ ID確認データをエクセルに保存成功: フォルダ={folder}")
                return True
                
            except Exception as e:
                logger.error(f"❌ ID確認エクセル保存処理中にエラー: {e}", exc_info=True)
                # 一時ファイルのクリーンアップ
                if temp_file and os.path.exists(temp_file.name):
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                return False
                
        except Exception as e:
            logger.error(f"❌ ID確認エクセル保存の初期化でエラー: {e}", exc_info=True)
            return False

def _verify_saved_data(excel_filepath: str, expected_folder: str, expected_id: str) -> bool:
    """保存されたデータの整合性を確認
    
    Args:
        excel_filepath: Excelファイルパス
        expected_folder: 期待されるフォルダ名
        expected_id: 期待されるID
        
    Returns:
        bool: 整合性確認結果
    """
    try:
        import openpyxl
        
        if not os.path.exists(excel_filepath):
            logger.error("📊 整合性確認: Excelファイルが存在しません")
            return False
        
        # Excelファイルを読み込み
        workbook = openpyxl.load_workbook(excel_filepath)
        worksheet = workbook.active
        
        # 最後の行を確認
        last_row = worksheet.max_row
        if last_row < 2:  # ヘッダー行しかない場合
            logger.error("📊 整合性確認: データ行が見つかりません")
            return False
        
        # 最後の行のデータを確認
        saved_folder = worksheet[f'A{last_row}'].value
        saved_id = worksheet[f'C{last_row}'].value
        
        # データが一致するかチェック
        folder_match = str(saved_folder) == str(expected_folder)
        id_match = str(saved_id) == str(expected_id)
        
        if folder_match and id_match:
            logger.info(f"📊 整合性確認OK: フォルダ={saved_folder}, ID={saved_id}")
            return True
        else:
            logger.error(f"📊 整合性確認NG: 期待値[{expected_folder}, {expected_id}] != 実際値[{saved_folder}, {saved_id}]")
            return False
            
    except Exception as e:
        logger.error(f"📊 整合性確認エラー: {e}")
        return False

def save_id_check_data_with_id(folder: str, id_image_path: str, copied_id: str) -> bool:
    """
    ID確認データをid_check.xlsxに保存します（ID数字を含む版）。
    
    Args:
        folder: フォルダ名
        id_image_path: ID部分の画像パス
        copied_id: クリップボードからコピーした数字ID
        
    Returns:
        bool: 保存成功かどうか
    """
    import openpyxl
    try:
        from openpyxl.drawing.image import Image as ExcelImage
    except ImportError:
        try:
            from openpyxl.drawing import Image as ExcelImage
        except ImportError:
            logger.error("openpyxlの画像機能が利用できません。画像埋め込みはスキップされます。")
            ExcelImage = None
    import tempfile
    import shutil
    
    # エクセルファイルパスを確定
    from utils.path_manager import get_base_path
    base_path = get_base_path()
    excel_filepath = os.path.join(base_path, "id_check.xlsx")
    
    # ファイルごとの排他制御ロック取得
    file_lock = _get_excel_lock(excel_filepath)
    
    with file_lock:  # 完全排他制御開始
        try:
            logger.info(f"📊 ID確認エクセルファイル保存開始: {excel_filepath} (フォルダ: {folder}, ID: {copied_id})")
            
            # 他の端末との競合を避けるため追加の待機時間
            base_delay = hash(folder) % 10 * 0.1  # フォルダに基づく0.0～0.9秒の遅延
            time.sleep(base_delay)
            
            # ディレクトリ作成
            os.makedirs(os.path.dirname(os.path.abspath(excel_filepath)) if os.path.dirname(excel_filepath) else '.', exist_ok=True)
        
            # 一時ファイルを使用した安全な書き込み
            temp_dir = os.path.dirname(excel_filepath) if os.path.dirname(excel_filepath) else '.'
            temp_file = None

            try:
                # エクセルファイルの読み込みまたは新規作成
                if os.path.exists(excel_filepath):
                    workbook = openpyxl.load_workbook(excel_filepath)
                    worksheet = workbook.active
                    logger.info("既存のid_check.xlsxファイルを読み込みました")
                else:
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    # ヘッダー作成（ID列を追加）
                    worksheet['A1'] = 'フォルダ名'
                    worksheet['B1'] = 'ID画像'  
                    worksheet['C1'] = 'コピーID'
                    worksheet['D1'] = '処理日時'
                    # 列幅設定
                    worksheet.column_dimensions['A'].width = 15
                    worksheet.column_dimensions['B'].width = 30
                    worksheet.column_dimensions['C'].width = 20
                    worksheet.column_dimensions['D'].width = 20
                    logger.info("新しいid_check.xlsxファイルを作成しました（ID列付き）")

                # 次の行番号を取得
                next_row = worksheet.max_row + 1
                
                # データを書き込み
                worksheet[f'A{next_row}'] = folder
                worksheet[f'C{next_row}'] = copied_id  # コピーしたID数字
                worksheet[f'D{next_row}'] = time.strftime("%Y-%m-%d %H:%M:%S")
                
                # ID画像を埋め込み
                if id_image_path and os.path.exists(id_image_path) and ExcelImage is not None:
                    try:
                        # 画像をエクセルに追加
                        img = ExcelImage(id_image_path)
                        # 画像サイズを調整（幅150px, 高さ30px程度）
                        img.width = 150
                        img.height = 30
                        # セルB列に画像を配置
                        img.anchor = f'B{next_row}'
                        worksheet.add_image(img)
                        
                        # 行の高さを画像に合わせて調整
                        worksheet.row_dimensions[next_row].height = 25
                        
                        logger.info(f"ID画像をエクセルに埋め込み: {id_image_path}")
                    except Exception as e:
                        logger.error(f"ID画像の埋め込みに失敗: {e}")
                        worksheet[f'B{next_row}'] = f"画像読み込み失敗: {os.path.basename(id_image_path)}"
                else:
                    # 画像埋め込みが利用できない場合、ファイルパスを記録
                    if id_image_path:
                        worksheet[f'B{next_row}'] = os.path.basename(id_image_path)
                        logger.info(f"ID画像パスを記録: {id_image_path}")
                    else:
                        worksheet[f'B{next_row}'] = "画像なし"
                        logger.warning("ID画像パスが無効です")
                
                # 一時ファイルに保存
                with tempfile.NamedTemporaryFile(mode='wb', delete=False, dir=temp_dir, suffix='.xlsx') as temp_file:
                    workbook.save(temp_file.name)
                    temp_filepath = temp_file.name
                
                # 一時ファイルを本来のファイルに移動（原子的操作）
                shutil.move(temp_filepath, excel_filepath)
                
                logger.info(f"✅ ID確認データをエクセルに保存成功: フォルダ={folder}, ID={copied_id}")
                
                # データの整合性を確認
                if _verify_saved_data(excel_filepath, folder, copied_id):
                    logger.info(f"📊 データ整合性確認完了: フォルダ={folder}, ID={copied_id}")
                    return True
                else:
                    logger.error(f"❌ データ整合性確認失敗: フォルダ={folder}, ID={copied_id}")
                    return False
                
            except Exception as e:
                logger.error(f"❌ ID確認エクセル保存処理中にエラー: {e}", exc_info=True)
                # 一時ファイルのクリーンアップ
                if temp_file and os.path.exists(temp_file.name):
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                return False
                
        except Exception as e:
            logger.error(f"❌ ID確認エクセル保存の初期化でエラー: {e}", exc_info=True)
            return False

def _verify_saved_data(excel_filepath: str, expected_folder: str, expected_id: str) -> bool:
    """保存されたデータの整合性を確認
    
    Args:
        excel_filepath: Excelファイルパス
        expected_folder: 期待されるフォルダ名
        expected_id: 期待されるID
        
    Returns:
        bool: 整合性確認結果
    """
    try:
        import openpyxl
        
        if not os.path.exists(excel_filepath):
            logger.error("📊 整合性確認: Excelファイルが存在しません")
            return False
        
        # Excelファイルを読み込み
        workbook = openpyxl.load_workbook(excel_filepath)
        worksheet = workbook.active
        
        # 最後の行を確認
        last_row = worksheet.max_row
        if last_row < 2:  # ヘッダー行しかない場合
            logger.error("📊 整合性確認: データ行が見つかりません")
            return False
        
        # 最後の行のデータを確認
        saved_folder = worksheet[f'A{last_row}'].value
        saved_id = worksheet[f'C{last_row}'].value
        
        # データが一致するかチェック
        folder_match = str(saved_folder) == str(expected_folder)
        id_match = str(saved_id) == str(expected_id)
        
        if folder_match and id_match:
            logger.info(f"📊 整合性確認OK: フォルダ={saved_folder}, ID={saved_id}")
            return True
        else:
            logger.error(f"📊 整合性確認NG: 期待値[{expected_folder}, {expected_id}] != 実際値[{saved_folder}, {saved_id}]")
            return False
            
    except Exception as e:
        logger.error(f"📊 整合性確認エラー: {e}")
        return False