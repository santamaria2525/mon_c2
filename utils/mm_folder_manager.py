"""
utils.mm_folder_manager - MMフォルダ管理機能

bin_pushフォルダ内のデータをMMフォルダ構造に変換・コピーする機能を提供します。
"""

from __future__ import annotations

import os
import shutil
from pathlib import Path
from typing import Dict, List, Tuple, Optional

from logging_util import logger
from .path_manager import get_base_path

class MMFolderManager:
    """MMフォルダの管理を行うクラス"""
    
    def __init__(self):
        self.base_path = get_base_path()
        self.bin_push_path = os.path.join(self.base_path, "bin_push")
        self.mm_folder_path = os.path.join(self.base_path, "MMフォルダ")
        
    def get_folder_mapping(self) -> Dict[str, Tuple[str, str]]:
        """フォルダ番号から(MM番号, 変換後フォルダ名)へのマッピングを取得
        
        Returns:
            Dict[str, Tuple[str, str]]: {元フォルダ番号: (MM番号, 変換後フォルダ名)}
            例: {"001": ("MM1", "001"), "1001": ("MM2", "001"), "2001": ("MM3", "001")}
        """
        mapping = {}
        
        # 1-999: そのままMM1へ
        for i in range(1, 1000):
            folder_num = f"{i:03d}"
            mapping[folder_num] = ("MM1", folder_num)
        
        # 1001-1999: -1000してMM2へ
        for i in range(1001, 2000):
            folder_num = f"{i:04d}"
            converted_num = f"{i-1000:03d}"
            mapping[folder_num] = ("MM2", converted_num)
        
        # 2001-2999: -2000してMM3へ
        for i in range(2001, 3000):
            folder_num = f"{i:04d}"
            converted_num = f"{i-2000:03d}"
            mapping[folder_num] = ("MM3", converted_num)
            
        return mapping
    
    def scan_bin_push_folders(self) -> List[str]:
        """bin_pushフォルダ内の既存フォルダをスキャン
        
        Returns:
            List[str]: 存在するフォルダ番号のリスト
        """
        if not os.path.exists(self.bin_push_path):
            logger.warning(f"bin_pushフォルダが見つかりません: {self.bin_push_path}")
            return []
        
        folders = []
        for item in os.listdir(self.bin_push_path):
            item_path = os.path.join(self.bin_push_path, item)
            if os.path.isdir(item_path) and item.isdigit():
                folders.append(item)
        
        folders.sort()
        return folders
    
    def copy_folder_to_mm(self, source_folder: str, mm_number: str, target_folder: str) -> bool:
        """指定フォルダをMMフォルダ構造にコピー
        
        Args:
            source_folder: bin_push内のソースフォルダ番号
            mm_number: 対象のMM番号 (MM1, MM2, MM3)
            target_folder: コピー先のフォルダ名
            
        Returns:
            bool: コピー成功時True
        """
        try:
            source_path = os.path.join(self.bin_push_path, source_folder)
            target_path = os.path.join(self.mm_folder_path, mm_number, target_folder)
            
            if not os.path.exists(source_path):
                logger.warning(f"ソースフォルダが存在しません: {source_path}")
                return False
            
            # ターゲットディレクトリを作成
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            
            # 既存のターゲットフォルダがあれば削除
            if os.path.exists(target_path):
                shutil.rmtree(target_path)
            
            # フォルダをコピー
            shutil.copytree(source_path, target_path)
            logger.debug(f"コピー完了: {source_folder} → {mm_number}/{target_folder}")
            return True
            
        except Exception as e:
            logger.error(f"フォルダコピー失敗: {source_folder} → {mm_number}/{target_folder}, エラー: {e}")
            return False
    
    def create_mm_folder_structure(self, target_ranges: Optional[List[str]] = None) -> Dict[str, int]:
        """MMフォルダ構造を作成
        
        Args:
            target_ranges: 対象範囲のリスト ["MM1", "MM2", "MM3"]。Noneの場合は全て
            
        Returns:
            Dict[str, int]: {MM番号: コピー成功数}
        """
        if target_ranges is None:
            target_ranges = ["MM1", "MM2", "MM3"]
        
        # bin_pushフォルダをスキャン
        existing_folders = self.scan_bin_push_folders()
        if not existing_folders:
            logger.error("bin_pushフォルダにデータが見つかりません")
            return {}
        
        # フォルダマッピングを取得
        folder_mapping = self.get_folder_mapping()
        
        # 統計情報
        stats = {mm: 0 for mm in target_ranges}
        total_processed = 0
        
        logger.info(f"● MMフォルダ作成開始: 対象範囲 {target_ranges}")
        
        for source_folder in existing_folders:
            if source_folder not in folder_mapping:
                continue
                
            mm_number, target_folder = folder_mapping[source_folder]
            
            if mm_number not in target_ranges:
                continue
            
            if self.copy_folder_to_mm(source_folder, mm_number, target_folder):
                stats[mm_number] += 1
                total_processed += 1
        
        # 結果報告
        logger.info(f"● MMフォルダ作成完了: 総処理数 {total_processed}")
        for mm_number in target_ranges:
            if stats[mm_number] > 0:
                logger.info(f"  - {mm_number}: {stats[mm_number]}フォルダ")
        
        return stats
    
    def get_mm_folder_info(self) -> Dict[str, Dict[str, int]]:
        """MMフォルダの現在の状況を取得
        
        Returns:
            Dict[str, Dict[str, int]]: MMフォルダの統計情報
        """
        info = {}
        
        if not os.path.exists(self.mm_folder_path):
            return info
        
        for mm_number in ["MM1", "MM2", "MM3"]:
            mm_path = os.path.join(self.mm_folder_path, mm_number)
            if os.path.exists(mm_path):
                folder_count = len([
                    item for item in os.listdir(mm_path)
                    if os.path.isdir(os.path.join(mm_path, item))
                ])
                info[mm_number] = {"folder_count": folder_count}
            else:
                info[mm_number] = {"folder_count": 0}
        
        return info
    
    def clean_mm_folder(self, mm_numbers: Optional[List[str]] = None) -> bool:
        """MMフォルダを清掃
        
        Args:
            mm_numbers: 清掃対象のMM番号リスト。Noneの場合は全て
            
        Returns:
            bool: 清掃成功時True
        """
        if mm_numbers is None:
            mm_numbers = ["MM1", "MM2", "MM3"]
        
        try:
            for mm_number in mm_numbers:
                mm_path = os.path.join(self.mm_folder_path, mm_number)
                if os.path.exists(mm_path):
                    shutil.rmtree(mm_path)
                    logger.info(f"● {mm_number}フォルダを削除しました")
            
            # MMフォルダ自体が空になったら削除
            if os.path.exists(self.mm_folder_path):
                remaining_items = os.listdir(self.mm_folder_path)
                if not remaining_items:
                    os.rmdir(self.mm_folder_path)
                    logger.info("● MMフォルダを削除しました")
            
            return True
            
        except Exception as e:
            logger.error(f"MMフォルダ清掃エラー: {e}")
            return False

    def batch_rename_folders(self, rename_mapping: Dict[str, str]) -> Dict[str, bool]:
        """フォルダの一括リネーム処理
        
        Args:
            rename_mapping: {旧フォルダ名: 新フォルダ名} のマッピング
            
        Returns:
            Dict[str, bool]: {フォルダ名: 成功/失敗フラグ}
        """
        results = {}
        
        if not os.path.exists(self.bin_push_path):
            logger.error(f"bin_pushフォルダが存在しません: {self.bin_push_path}")
            return results
        
        # 一時的なバックアップディレクトリを作成
        backup_path = os.path.join(self.base_path, "rename_backup")
        target_path = os.path.join(self.base_path, "rename_result")
        
        try:
            # バックアップディレクトリとターゲットディレクトリを作成
            os.makedirs(backup_path, exist_ok=True)
            os.makedirs(target_path, exist_ok=True)
            
            # 既存のターゲットフォルダをクリア
            if os.path.exists(target_path):
                for item in os.listdir(target_path):
                    item_path = os.path.join(target_path, item)
                    if os.path.isdir(item_path):
                        shutil.rmtree(item_path)
            
            success_count = 0
            
            for old_name, new_name in rename_mapping.items():
                try:
                    source_path = os.path.join(self.bin_push_path, old_name)
                    target_full_path = os.path.join(target_path, new_name)
                    
                    if not os.path.exists(source_path):
                        logger.warning(f"リネーム対象フォルダが見つかりません: {old_name}")
                        results[old_name] = False
                        continue
                    
                    if not os.path.isdir(source_path):
                        logger.warning(f"リネーム対象がフォルダではありません: {old_name}")
                        results[old_name] = False
                        continue
                    
                    # フォルダをコピー（リネーム）
                    shutil.copytree(source_path, target_full_path)
                    results[old_name] = True
                    success_count += 1
                    logger.info(f"フォルダリネーム成功: '{old_name}' → '{new_name}'")
                    
                except Exception as e:
                    logger.error(f"フォルダリネーム失敗: '{old_name}' → '{new_name}', エラー: {e}")
                    results[old_name] = False
            
            logger.info(f"● フォルダ一括リネーム完了: {success_count}/{len(rename_mapping)} 成功")
            return results
            
        except Exception as e:
            logger.error(f"フォルダ一括リネーム処理でエラー発生: {e}")
            return results

    def load_excel_mapping(self, excel_path: str) -> Optional[Dict[str, str]]:
        """Excelファイルからリネームマッピングを読み込み
        
        Args:
            excel_path: Excelファイルのパス
            
        Returns:
            Dict[str, str]: {旧フォルダ名: 新フォルダ名} または None
        """
        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excelファイルが見つかりません: {excel_path}")
                return None
            
            # openpyxlを使ってExcelファイルを読み込み
            from openpyxl import load_workbook
            
            workbook = load_workbook(excel_path, data_only=True)
            worksheet = workbook.active
            
            mapping = {}
            
            # A列: 旧フォルダ名, B列: 新フォルダ名として読み込み
            for row_num, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if len(row) >= 2:
                    old_name = row[0]
                    new_name = row[1]
                    
                    # None値をスキップ
                    if old_name is None or new_name is None:
                        continue
                        
                    # 数値を文字列に変換（0パディング考慮）
                    if isinstance(old_name, (int, float)):
                        old_name = f"{int(old_name):03d}" if old_name < 1000 else f"{int(old_name):04d}"
                    else:
                        old_name = str(old_name).strip()
                        
                    if isinstance(new_name, (int, float)):
                        new_name = f"{int(new_name):03d}" if new_name < 1000 else f"{int(new_name):04d}"
                    else:
                        new_name = str(new_name).strip()
                    
                    if old_name and new_name:
                        mapping[old_name] = new_name
                    else:
                        logger.warning(f"Excelの{row_num}行目: 空の値があります")
                else:
                    if any(cell is not None for cell in row[:2] if row):
                        logger.warning(f"Excelの{row_num}行目: 列数が不足しています")
            
            workbook.close()
            logger.info(f"Excelマッピング読み込み完了: {len(mapping)}件")
            return mapping
            
        except Exception as e:
            logger.error(f"Excelファイル読み込みエラー: {e}")
            return None
    
    def load_csv_mapping(self, csv_path: str) -> Optional[Dict[str, str]]:
        """CSVファイルからリネームマッピングを読み込み（互換性維持用）
        
        Args:
            csv_path: CSVファイルのパス
            
        Returns:
            Dict[str, str]: {旧フォルダ名: 新フォルダ名} または None
        """
        import csv
        
        try:
            if not os.path.exists(csv_path):
                logger.error(f"CSVファイルが見つかりません: {csv_path}")
                return None
            
            mapping = {}
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                for row_num, row in enumerate(reader, 1):
                    if len(row) >= 2:
                        old_name = row[0].strip()
                        new_name = row[1].strip()
                        if old_name and new_name:
                            mapping[old_name] = new_name
                        else:
                            logger.warning(f"CSVの{row_num}行目: 空の値があります")
                    else:
                        logger.warning(f"CSVの{row_num}行目: 列数が不足しています")
            
            logger.info(f"CSVマッピング読み込み完了: {len(mapping)}件")
            return mapping
            
        except Exception as e:
            logger.error(f"CSVファイル読み込みエラー: {e}")
            return None

# 便利関数
def create_mm_folders(target_ranges: Optional[List[str]] = None) -> Dict[str, int]:
    """MMフォルダを作成する便利関数
    
    Args:
        target_ranges: 対象範囲 ["MM1", "MM2", "MM3"]
        
    Returns:
        Dict[str, int]: 作成統計
    """
    manager = MMFolderManager()
    return manager.create_mm_folder_structure(target_ranges)

def get_mm_folder_status() -> Dict[str, Dict[str, int]]:
    """MMフォルダの状況を取得する便利関数
    
    Returns:
        Dict[str, Dict[str, int]]: MMフォルダ統計
    """
    manager = MMFolderManager()
    return manager.get_mm_folder_info()

def clean_mm_folders(mm_numbers: Optional[List[str]] = None) -> bool:
    """MMフォルダを清掃する便利関数
    
    Args:
        mm_numbers: 清掃対象のMM番号
        
    Returns:
        bool: 清掃成功時True
    """
    manager = MMFolderManager()
    return manager.clean_mm_folder(mm_numbers)

def batch_rename_folders_excel(excel_path: str) -> Dict[str, bool]:
    """Excelファイルを使ってフォルダを一括リネームする便利関数
    
    Args:
        excel_path: Excelファイルのパス（A列:旧フォルダ名, B列:新フォルダ名の形式）
        
    Returns:
        Dict[str, bool]: {フォルダ名: 成功/失敗フラグ}
    """
    manager = MMFolderManager()
    mapping = manager.load_excel_mapping(excel_path)
    if mapping is None:
        return {}
    return manager.batch_rename_folders(mapping)

def batch_rename_folders_csv(csv_path: str) -> Dict[str, bool]:
    """CSVファイルを使ってフォルダを一括リネームする便利関数（互換性維持用）
    
    Args:
        csv_path: CSVファイルのパス（旧フォルダ名,新フォルダ名の形式）
        
    Returns:
        Dict[str, bool]: {フォルダ名: 成功/失敗フラグ}
    """
    manager = MMFolderManager()
    mapping = manager.load_csv_mapping(csv_path)
    if mapping is None:
        return {}
    return manager.batch_rename_folders(mapping)