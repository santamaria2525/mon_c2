# -*- coding: utf-8 -*-
"""Excel-driven account backup executor."""

from __future__ import annotations

import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Optional, Sequence, Mapping, Tuple

import openpyxl  # type: ignore

from adb_utils import reset_adb_server
from logging_util import MultiDeviceLogger, logger
from utils import display_message, get_target_folder
from config import get_config_value

from services import ConfigService, MultiDeviceService
from missing_functions import device_operation_excel_and_save


class AccountBackupExecutor:
    """Handle the Excel-based account backup workflow."""

    def __init__(
        self,
        core,
        config_service: ConfigService,
        multi_device_service: MultiDeviceService,
    ) -> None:
        self.core = core
        self.config_service = config_service
        self.multi_device_service = multi_device_service

    def run(self) -> None:
        ports = self._resolve_ports()
        if not ports:
            return

        workbook = self._load_workbook("mon_aco.xlsx")
        if workbook is None:
            return

        sheet = workbook.active
        total_rows = sheet.max_row

        current_row = self._resolve_start_row(total_rows)
        if current_row is None:
            return

        logger.info("Excelバックアップ開始 ports=%s start_row=%d", ports, current_row)
        reset_adb_server()
        skip_nox_restart = bool(get_config_value("skip_nox_restart_on_fail", False))

        # Portごとの進捗を管理し、空いた端末から順に次の行を処理する
        pending_rows = list(range(current_row, total_rows + 1))
        completed_rows: set[int] = set()
        active_rows: dict[str, int] = {}

        while pending_rows or active_rows:
            if self.core.is_stopping():
                break

            # 空いているポートがあれば次の行を割り当てる
            for port in ports:
                if port in active_rows:
                    continue
                while pending_rows and pending_rows[0] in completed_rows:
                    pending_rows.pop(0)
                if not pending_rows:
                    break
                row = pending_rows.pop(0)
                if row > total_rows:
                    continue
                if not skip_nox_restart:
                    self.multi_device_service.remove_all_nox([port])
                logger.debug("バックアップ開始: port=%s row=%s", port, row)
                active_rows[port] = row

            if not active_rows:
                break

            results = self._process_active_rows(workbook, active_rows)
            for port, result in results.items():
                row = active_rows.pop(port, None)
                if row is None:
                    continue
                row_label = f"{row:03d}"
                success, error_message = result
                if success:
                    completed_rows.add(row)
                    logger.info(
                        "Excelバックアップ: %s行、成功 (port=%s) / 累計完了=%d行",
                        row_label,
                        port,
                        len(completed_rows),
                    )
                else:
                    reason = error_message or "原因不明のエラー"
                    logger.error(
                        "Excelバックアップ: %s行、%sにより失敗 (port=%s)",
                        row_label,
                        reason,
                        port,
                    )
                    logger.debug(
                        "Excelバックアップ: 行%s 失敗詳細 (port=%s) -> %s",
                        row_label,
                        port,
                        reason,
                    )
                    should_restart = (
                        not skip_nox_restart
                        and (not reason or any(keyword in reason for keyword in ("タイムアップ", "フリーズ", "応答なし")))
                    )
                    if should_restart:
                        self.multi_device_service.remove_all_nox([port])
                        logger.warning("Excelバックアップ: port=%s のNOXを再起動しました", port)
                    pending_rows.append(row)

        logger.info("Excelバックアップ完了")

    # ------------------------------------------------------------------ #
    # Helpers
    # ------------------------------------------------------------------ #
    def _resolve_ports(self) -> Sequence[str]:
        snapshot = self.config_service.load()
        ports = self.config_service.get_ports_for_device_count(snapshot.device_count)
        if not ports:
            logger.error("Excelバックアップ: ポート設定が見つかりません。config.json を確認してください。")
        return ports

    def _load_workbook(self, filename: str):
        path = Path(filename)
        if not path.exists():
            logger.error("Excelバックアップ: %s が見つかりません。", filename)
            display_message("エラー", f"{filename} が見つかりません。")
            return None
        try:
            return openpyxl.load_workbook(str(path))
        except Exception as exc:
            logger.error("Excelバックアップ: %s の読み込みに失敗しました (%s)", filename, exc)
            display_message("エラー", f"{filename} の読み込みに失敗しました。\n{exc}")
            return None

    def _resolve_start_row(self, total_rows: int) -> Optional[int]:
        row = get_target_folder()
        if row is None:
            logger.info("Excelバックアップ: 行番号が指定されなかったため中断します。")
            return None
        try:
            value = int(row)
        except ValueError:
            logger.error("Excelバックアップ: 無効な行番号 %s", row)
            display_message("エラー", "行番号は整数で入力してください。")
            return None
        if value < 1 or value > total_rows:
            logger.error("Excelバックアップ: 行番号 %d は範囲外です (1-%d)", value, total_rows)
            display_message("エラー", f"行番号は 1 から {total_rows} の範囲で指定してください。")
            return None
        return value

    def _process_active_rows(
        self, workbook, active_rows: Mapping[str, int]
    ) -> Mapping[str, tuple[bool, str]]:
        """処理中の行を並列実行し、ポートごとの成否とエラーメッセージを返す。"""
        ports = list(active_rows.keys())
        folders = [f"{row:03d}" for row in active_rows.values()]
        ml = MultiDeviceLogger(ports, folders)
        results: dict[str, bool] = {}

        with ThreadPoolExecutor(max_workers=len(ports)) as executor:
            future_map = {
                executor.submit(
                    device_operation_excel_and_save,
                    port,
                    workbook,
                    row,
                    row,
                    threading.Event(),
                    ml,
                ): port
                for port, row in active_rows.items()
            }

            for future in as_completed(future_map):
                port = future_map[future]
                try:
                    results[port] = bool(future.result())
                except Exception as exc:  # pragma: no cover - defensive logging
                    logger.error("Excelバックアップ: %s で例外が発生しました: %s", port, exc)
                    results[port] = False

        ml.summarize_results("Excelバックアップ", suppress_summary=True)
        detailed_results: dict[str, tuple[bool, str]] = {}
        for port, ok in results.items():
            detailed_results[port] = (ok, ml.get_error(port))
        return detailed_results


__all__ = ["AccountBackupExecutor"]
