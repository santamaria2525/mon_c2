"""
monst.device.checks - Character and progress checking operations.

キャラクター確認や進行状況チェック機能を提供します。
"""

from __future__ import annotations

import os
import time
import threading
from datetime import datetime
from typing import Optional

from config import on_check
from logging_util import logger, MultiDeviceLogger
from monst.adb import perform_action
from monst.image import tap_if_found, tap_until_found
from utils.path_manager import get_base_path

from .navigation import home
from openpyxl import Workbook, load_workbook

_kuribo_excel_lock = threading.Lock()

def icon_check(
    device_port: str, 
    folder: str, 
    multi_logger: Optional[MultiDeviceLogger] = None
) -> None:
    """各種チェック機能を実行します。
    
    設定に応じて以下のチェックを実行：
    - ノマクエチェック (on_check=1)
    - 覇者の塔クリアチェック (on_check=2)  
    - 守護獣所持チェック (on_check=3)
    - クリボー確認 (on_check=4)
    
    Args:
        device_port: 対象デバイスのポート
        folder: フォルダ名
        multi_logger: マルチデバイスロガー（オプション）
        
    Example:
        >>> icon_check("127.0.0.1:62001", "folder_001")
    """
    if not home(device_port, folder):
        logger.warning(f"デバイス {device_port}: home関数が失敗しましたが、処理を継続します")

    if on_check == 1:  # ノマクエチェック
        _check_normal_quest(device_port, folder)
    elif on_check == 2:  # 覇者の塔クリアチェック
        _check_tower_completion(device_port, folder)
    elif on_check == 3:  # 守護獣所持チェック
        _check_guardian_beasts(device_port, folder)
    elif on_check == 4:  # クリボー確認
        _check_kuribo(device_port, folder)

def _check_normal_quest(device_port: str, folder: str) -> None:
    """ノマクエの進行状況をチェックします。"""
    max_attempts = 10
    for attempt in range(max_attempts):
        if tap_if_found('tap', device_port, "noma.png", "key"):
            break
        tap_if_found('tap', device_port, "quest_c.png", "key")
        tap_if_found('tap', device_port, "quest.png", "key")
        tap_if_found('tap', device_port, "ichiran.png", "key")
        tap_if_found('tap', device_port, "ok.png", "key")
        tap_if_found('tap', device_port, "close.png", "key")
        time.sleep(1)

def _check_tower_completion(device_port: str, folder: str) -> None:
    """覇者の塔のクリア状況をチェックします。"""
    timeout: int = 60  # タイムアウト時間（秒単位）
    start_time: float = time.time()  # 処理開始時間を記録

    while True:
        # "hasyafin1.png" を探す
        if tap_if_found('tap', device_port, "hasyafin1.png", "key"):
            break  # 見つかった場合、ループを終了

        # 2分間見つからなかった場合
        if time.time() - start_time > timeout:
            logger.info(f"覇者未完了。対象フォルダ: {folder}")
            break

        # 他のタップ処理を実行
        tap_if_found('tap', device_port, "quest_c.png", "key")
        tap_if_found('tap', device_port, "quest.png", "key")
        tap_if_found('tap', device_port, "ichiran.png", "key")
        tap_if_found('tap', device_port, "ok.png", "key")
        tap_if_found('tap', device_port, "close.png", "key")
        time.sleep(1)  # 次のチェックまで待機

def _check_guardian_beasts(device_port: str, folder: str) -> None:
    """守護獣の所持状況をチェックします（mon6準拠）。"""
    tap_until_found(device_port, "monbox.png", "key", "monster.png", "key", "tap")
    tap_until_found(device_port, "shugo_box2.png", "key", "shugo_box.png", "key", "tap")
    tap_until_found(device_port, "shugo_ishi.png", "key", "ok.png", "key", "tap")
    
    # mon6準拠の守護獣所持チェック
    if not tap_if_found('stay', device_port, "shugo1.png", "icon"):
        logger.info(f"守護１未所持　対象フォルダ: {folder}")
    if not tap_if_found('stay', device_port, "shugo2.png", "icon"):
        logger.info(f"守護２未所持　対象フォルダ: {folder}")
    if not tap_if_found('stay', device_port, "shugo3.png", "icon"):
        logger.info(f"守護３未所持　対象フォルダ: {folder}")
    if not tap_if_found('stay', device_port, "shugo4.png", "icon"):
        logger.info(f"守護４未所持　対象フォルダ: {folder}")


def _check_kuribo(device_port: str, folder: str) -> None:
    """クリボー確認（on_check=4）。"""
    # ② sort.pngが表示されるまでUI遷移を繰り返す
    navigation_images = [
        "monster.png",
        "monster_box.png",
        "ok.png",
        "close.png",
        "hyoujijun.png",
    ]
    max_attempts = 20
    for attempt in range(max_attempts):
        if tap_if_found('stay', device_port, "sort.png", "ui"):
            break
        for image_name in navigation_images:
            tap_if_found('tap', device_port, image_name, "ui")
            time.sleep(0.3)
    else:
        logger.warning(f"sort.pngが見つからずクリボー確認をスキップします: フォルダ {folder}")
        _record_kuribo_result(folder, has_kuribo=False, detected=False)
        return

    # ③ ソート条件を選択
    tap_if_found('tap', device_port, "z_hoshi3.png", "ui")
    tap_if_found('tap', device_port, "z_yami.png", "ui")
    tap_if_found('tap', device_port, "kettei_mon.png", "ui")

    # ④ 2秒待機してクリボーを検出
    time.sleep(2.0)
    has_kuribo = tap_if_found('stay', device_port, "kuribo.png", "ui")
    if has_kuribo:
        logger.info(f"クリボーあり: フォルダ {folder}")
    else:
        logger.info(f"クリボーなし: フォルダ {folder}")
    _record_kuribo_result(folder, has_kuribo=has_kuribo, detected=True)


def _record_kuribo_result(folder: str, has_kuribo: bool, *, detected: bool) -> None:
    """クリボー確認結果をExcelに記録する。"""
    try:
        base_path = get_base_path()
        excel_path = os.path.join(base_path, "kuribo_check.xlsx")
        os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)

        with _kuribo_excel_lock:
            if os.path.exists(excel_path):
                workbook = load_workbook(excel_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "kuribo"
                sheet.append(["Timestamp", "Folder", "Result", "Detected"])

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            result_label = "あり" if has_kuribo else "なし"
            sheet.append([timestamp, folder, result_label, "Yes" if detected else "No"])
            workbook.save(excel_path)
            logger.debug(f"クリボー確認をExcel保存: {excel_path} ({folder}: {result_label})")
    except Exception as exc:
        logger.error(f"クリボー確認結果の保存に失敗しました: {exc}")
